from pathlib import Path
from threading import Lock
from datetime import datetime
import hashlib
import secrets
import json

import joblib
import numpy as np
from flask import Flask, jsonify, request, session
from flask_cors import CORS
from tensorflow.keras.models import load_model
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


app = Flask(__name__)
CORS(app)
app.secret_key = secrets.token_hex(32)

MODEL_DIR = Path(__file__).resolve().parent / "models"
DATA_DIR = Path(__file__).resolve().parent / "data"
DATA_DIR.mkdir(exist_ok=True)

PATIENT_DB_FILE = DATA_DIR / "patient_database.xlsx"

MODEL_LOCK = Lock()
MODEL = None
SCALER = None
TARGET_ENCODER = None
FEATURE_COLUMNS = None
RISK_LEVEL_MAP = None


def init_excel_workbook():
    """Initialize Excel workbook with sheets if it doesn't exist."""
    if not PATIENT_DB_FILE.exists():
        wb = Workbook()
        
        # Create Users sheet
        ws_users = wb.active
        ws_users.title = "Users"
        headers_users = ['user_id', 'name', 'password_hash', 'created_at']
        ws_users.append(headers_users)
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_users[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Create Patient Data sheet
        ws_patients = wb.create_sheet("Patient Data")
        headers_patients = ['patient_id', 'user_id', 'user_name', 'gender', 'age', 'height', 'weight', 
                           'family_history_with_overweight', 'FAVC', 'FCVC', 'NCP', 'CAEC', 'CH2O', 
                           'SCC', 'PAL', 'MTRANS', 'predicted_risk_level', 'confidence', 'created_at', 'SMOKE', 'FAF', 'TUE', 'CALC']
        ws_patients.append(headers_patients)
        
        # Style header row
        for cell in ws_patients[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for ws in [ws_users, ws_patients]:
            for column in ws.columns:
                max_length = 12
                column_letter = column[0].column_letter
                ws.column_dimensions[column_letter].width = max_length
        
        wb.save(PATIENT_DB_FILE)
        return wb
    else:
        return load_workbook(PATIENT_DB_FILE)


def hash_password(password):
    """Hash a password using SHA-256."""
    return hashlib.sha256(password.encode()).hexdigest()


def get_next_user_id():
    """Get the next available user ID."""
    wb = load_workbook(PATIENT_DB_FILE)
    ws = wb["Users"]
    return ws.max_row  # max_row includes header, so this is correct


def get_next_patient_id():
    """Get the next available patient ID."""
    wb = load_workbook(PATIENT_DB_FILE)
    ws = wb["Patient Data"]
    return ws.max_row  # max_row includes header


def user_exists(name):
    """Check if a username already exists."""
    wb = load_workbook(PATIENT_DB_FILE)
    ws = wb["Users"]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == name:  # name is in column 2 (index 1)
            return True
    return False


def get_user_by_name(name):
    """Get user by name."""
    wb = load_workbook(PATIENT_DB_FILE)
    ws = wb["Users"]
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[1].value == name:  # name is in column 2
            return {
                'user_id': row[0].value,
                'name': row[1].value,
                'password_hash': row[2].value,
                'created_at': row[3].value
            }
    return None


def load_artifacts() -> None:
    global MODEL, SCALER, TARGET_ENCODER, FEATURE_COLUMNS, RISK_LEVEL_MAP

    if MODEL is not None:
        return

    with MODEL_LOCK:
        if MODEL is None:
            MODEL = load_model(MODEL_DIR / "ann_obesity_model.keras")
            SCALER = joblib.load(MODEL_DIR / "scaler.pkl")
            TARGET_ENCODER = joblib.load(MODEL_DIR / "target_encoder.pkl")
            FEATURE_COLUMNS = joblib.load(MODEL_DIR / "feature_columns.pkl")
            RISK_LEVEL_MAP = joblib.load(MODEL_DIR / "risk_level_map.pkl")


def parse_payload(payload: dict) -> np.ndarray:
    missing_fields = [key for key in FEATURE_COLUMNS if key not in payload]
    if missing_fields:
        raise ValueError(f"Missing required fields: {missing_fields}")

    values = []
    for feature in FEATURE_COLUMNS:
        raw_value = payload[feature]
        try:
            values.append(float(raw_value))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid value for field '{feature}': {raw_value}") from exc

    return np.array(values, dtype=np.float32).reshape(1, -1)


@app.get("/")
def home():
    return jsonify(
        {
            "service": "ObesiCare API",
            "status": "running",
            "endpoint": "/predict",
        }
    )


@app.post("/auth/register")
def register():
    """Register a new user with name and password."""
    try:
        init_excel_workbook()
        
        payload = request.get_json(silent=True)
        if not payload:
            return jsonify({"error": "Request body must be valid JSON."}), 400
        
        name = payload.get('name', '').strip()
        password = payload.get('password', '').strip()
        
        if not name or len(name) < 2:
            return jsonify({"error": "Name must be at least 2 characters."}), 400
        
        if not password or len(password) < 6:
            return jsonify({"error": "Password must be at least 6 characters."}), 400
        
        if user_exists(name):
            return jsonify({"error": "This username already exists."}), 409
        
        user_id = get_next_user_id()
        password_hash = hash_password(password)
        created_at = datetime.now().isoformat()
        
        wb = load_workbook(PATIENT_DB_FILE)
        ws = wb["Users"]
        ws.append([user_id, name, password_hash, created_at])
        wb.save(PATIENT_DB_FILE)
        
        return jsonify({
            "success": True,
            "message": "User registered successfully.",
            "user_id": user_id,
            "name": name
        }), 201
    
    except Exception as exc:
        return jsonify({"error": f"Registration failed: {exc}"}), 500


@app.post("/auth/login")
def login():
    """Login user with name and password."""
    try:
        init_excel_workbook()
        
        payload = request.get_json(silent=True)
        if not payload:
            return jsonify({"error": "Request body must be valid JSON."}), 400
        
        name = payload.get('name', '').strip()
        password = payload.get('password', '').strip()
        
        if not name or not password:
            return jsonify({"error": "Username and password are required."}), 400
        
        user = get_user_by_name(name)
        if not user:
            return jsonify({"error": "Invalid username or password."}), 401
        
        password_hash = hash_password(password)
        if user['password_hash'] != password_hash:
            return jsonify({"error": "Invalid username or password."}), 401
        
        session['user_id'] = user['user_id']
        session['user_name'] = user['name']
        
        return jsonify({
            "success": True,
            "message": "Login successful.",
            "user_id": user['user_id'],
            "name": user['name']
        }), 200
    
    except Exception as exc:
        return jsonify({"error": f"Login failed: {exc}"}), 500


@app.post("/auth/logout")
def logout():
    """Logout current user."""
    session.clear()
    return jsonify({"success": True, "message": "Logged out successfully."}), 200


@app.post("/predict")
def predict():
    try:
        init_excel_workbook()
        load_artifacts()

        payload = request.get_json(silent=True)
        if not payload:
            return jsonify({"error": "Request body must be valid JSON."}), 400

        # Extract user info
        user_id = payload.get('user_id')
        user_name = payload.get('user_name')
        
        if not user_id or not user_name:
            return jsonify({"error": "User ID and name are required for prediction."}), 400

        # Extract health features
        input_data = parse_payload(payload)
        input_scaled = SCALER.transform(input_data)

        probabilities = MODEL.predict(input_scaled, verbose=0)[0]
        predicted_class_index = int(np.argmax(probabilities))
        confidence = float(probabilities[predicted_class_index])

        predicted_risk_level = RISK_LEVEL_MAP.get(
            predicted_class_index,
            str(TARGET_ENCODER.inverse_transform([predicted_class_index])[0]),
        )

        # Store patient data
        patient_id = get_next_patient_id()
        created_at = datetime.now().isoformat()
        
        patient_record = [
            patient_id,
            user_id,
            user_name,
            payload.get('Gender', ''),
            payload.get('Age', ''),
            payload.get('Height', ''),
            payload.get('Weight', ''),
            payload.get('family_history_with_overweight', ''),
            payload.get('FAVC', ''),
            payload.get('FCVC', ''),
            payload.get('NCP', ''),
            payload.get('CAEC', ''),
            payload.get('CH2O', ''),
            payload.get('SCC', ''),
            payload.get('PAL', ''),
            payload.get('MTRANS', ''),
            predicted_risk_level,
            round(confidence, 4),
            created_at,
            payload.get('SMOKE', ''),
            payload.get('FAF', ''),
            payload.get('TUE', ''),
            payload.get('CALC', '')
        ]
        
        wb = load_workbook(PATIENT_DB_FILE)
        ws = wb["Patient Data"]
        ws.append(patient_record)
        
        # Color code risk level
        last_row = ws.max_row
        risk_cell = ws.cell(row=last_row, column=17)  # predicted_risk_level column
        
        if predicted_risk_level == "Insufficient_Weight":
            risk_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif predicted_risk_level in ["Normal_Weight", "Overweight_Level_I"]:
            risk_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif predicted_risk_level in ["Overweight_Level_II", "Obesity_Type_I"]:
            risk_cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        elif predicted_risk_level in ["Obesity_Type_II", "Obesity_Type_III"]:
            risk_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        wb.save(PATIENT_DB_FILE)

        return jsonify(
            {
                "predicted_class_index": predicted_class_index,
                "predicted_risk_level": predicted_risk_level,
                "confidence": round(confidence, 4),
                "patient_id": patient_id,
                "message": "Prediction saved to patient record in Excel."
            }
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except FileNotFoundError as exc:
        return jsonify({"error": f"Model artifacts not found: {exc}"}), 500
    except Exception as exc:
        return jsonify({"error": f"Unexpected server error: {exc}"}), 500


@app.get("/patient-history/<int:user_id>")
def get_patient_history(user_id):
    """Retrieve all health records for a patient (for personalization)."""
    try:
        init_excel_workbook()
        
        records = []
        wb = load_workbook(PATIENT_DB_FILE)
        ws = wb["Patient Data"]
        
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id:  # user_id is in column 2
                record = dict(zip(headers, row))
                records.append(record)
        
        if not records:
            return jsonify({"error": "No records found for this user."}), 404
        
        return jsonify({
            "user_id": user_id,
            "total_records": len(records),
            "records": records
        }), 200
    
    except Exception as exc:
        return jsonify({"error": f"Error retrieving history: {exc}"}), 500


if __name__ == "__main__":
    init_excel_workbook()
    app.run(debug=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
